import pandas as pd
import numpy as np
from typing import Optional
from sqlalchemy.orm import Session
from backend.models.database import FinancialDocument, Transaction, DocumentChunk, ExtractedTransactions
from backend.config.currency_utils_config import CurrencyDetector, detect_document_currency
from backend.services.batch_processor import BatchProcessor
from backend.db.redis_client import RedisCache, cached
from concurrent.futures import ThreadPoolExecutor, as_completed
from backend.features_enginering.features import categorize_transaction
from typing import List, Tuple, Dict, Callable
import hashlib
from datetime import datetime, timedelta
import os
import openpyxl
import traceback
from collections import Counter
import logging
import re

logger = logging.getLogger(__name__)

class ExcelCurrencyReader:
    """Reading excel file format for currency detector"""
    
    EXCEL_CURRENCY_MAP = {
        'Rp': 'IDR',
        'rp': 'IDR',
        'IDR': 'IDR',
        '$': 'USD',
        'USD': 'USD',
    }
    
    def detect_currency_from_format(self, number_format: str) -> Optional[str]:
        """Extract currency from excel number format string"""
        if not number_format or number_format == 'General':
            return None
        
        # check symbol inside double-qoutes
        quoted = re.findall(r'"([^"]+)"', number_format)
        for sym in quoted:
            sym_clean = sym.strip()
            for key, code in self.EXCEL_CURRENCY_MAP.items():
                if key.lower() == sym_clean.lower():
                    logger.debug(f"Currency '{code}' from quoted format symbol '{sym_clean}")
                    return code
                
        # fallback methods if above method fail
        for key, code in self.EXCEL_CURRENCY_MAP.items():
            if key in number_format:
                logger.debug(f"Currency '{code}' from raw format symbol '{key}'")
                return code
        
        return None
    
    def detect_currency_from_excel(self, file_path: str, amount_column: str = 'Amount') -> Optional[str]:
        """detect currency from excel cell formatting"""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=False)
            worksheet = workbook.active
            
            # find amount column index
            header_row = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            amount_col_idx = None
            
            for idx, cell_value in enumerate(header_row, start=1):
                if str(cell_value).strip().lower() == amount_column.strip().lower():
                    amount_col_idx = idx
                    break
                
            if amount_col_idx is None:
                logger.warning(f"Column '{amount_column}' not found in Excel headers")
                workbook.close()
                return None
            
            # get column letter
            col_letter = openpyxl.utils.get_column_letter(amount_col_idx)
            
            # sample the first 10 data rows to detect currency format
            detected = []
            for row_idx in range(2, min(12, worksheet.max_row + 1)):
                cell = worksheet[f'{col_letter}{row_idx}']
                
                if cell.value is not None:
                    currency = self.detect_currency_from_format(cell.number_format)
                    if currency:
                        detected.append(currency)
            
            workbook.close()
            
            # determine predominant currency
            if detected:
                most_common = Counter(detected).most_common(1)[0][0]
                logger.debug(f"Currency detected from Excel cell formatting: {most_common}")
                return most_common
            else:
                logger.warning("No currency detected from Excel formatting")
                return None
        except Exception as e:
            logger.error(f"❌ Error reading Excel formatting: {e}")
            return None
        
    def read_with_currency_formats(self, file_path: str, amount_column: str) -> dict:
        """Read the amount column and return per-index cells into currency map through loop"""
        row_currency_map = {}
        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=False)
            worksheet = workbook.active
            
            header_row = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            
            amount_col_idx = None
            for idx, cell_value in enumerate(header_row, start=1):
                if str(cell_value).strip().lower() == amount_column.strip().lower():
                    amount_col_idx = idx
                    break
                
            if amount_col_idx is None:
                workbook.close()
                return row_currency_map
            
            col_letter = openpyxl.utils.get_column_letter(amount_col_idx)
            
            # map pandas row index(based 0 in data rows) into currency
            for row_idx in range(2, worksheet.max_row + 1):
                cell = worksheet[f'{col_letter}{row_idx}']
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    currency = self.detect_currency_from_format(cell.number_format)
                    if currency:
                        pandas_idx = row_idx - 2
                        row_currency_map[pandas_idx] = currency
                        
            workbook.close()
            logger.info(f"Built per-row currency map: {len(row_currency_map)} numeric cells with known currency format")
            
        except Exception as e:
            logger.error(f"Error building row currency map: {e}")
        
        return row_currency_map

class DocumentService:
    def __init__(self, db: Session, user_currency: str = None):
        self.db = db
        self.user_currency = user_currency
        self.excel_reader = ExcelCurrencyReader()
        self.currency_detector = CurrencyDetector(base_currency='USD')
        self.detected_document_currency = user_currency or 'USD'
        
    def detect_file_currency(self, file_path: str, df: pd.DataFrame, amount_col: str) -> Optional[str]:
        """Smart currency detection from file format"""
        detected_currency = None
        
        # for excel files
        if file_path.endswith('.xlsx') or file_path.endswith('.xls'):
            logger.info("Excel file detected - checking cell formatting...")
            detected_currency = self.excel_reader.detect_currency_from_excel(file_path, amount_col)
            
            if detected_currency:
                logger.info(f"Currency from Excel format: {detected_currency}")
                return detected_currency
            
        # fallback to user preferences currency
        if self.user_currency:
            logger.info(f"Using user preference currency: {self.user_currency}")
            return self.user_currency
        
        logger.warning("No currency detected, defaulting to USD")
        return 'USD'
            
    def _read_dataframe(self, file_path: str) -> pd.DataFrame:
        """Method to unify to read CSV or Excel files"""
        try:
            print(f"📖 Reading file: {file_path}")

            if file_path.endswith('.csv'):
                # Try different encodings for CSV
                for encoding in ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252']:
                    try:
                        df = pd.read_csv(file_path, encoding=encoding)
                        print(f"✅ CSV read with {encoding} encoding")
                        
                        df.columns = [c.strip('""').strip("'") for c in df.columns]
                        for col in df.select_dtypes(include=['object', 'str']).columns:
                            df[col] = df[col].apply(
                                lambda x: x.strip('"').strip("'") if isinstance(x, str) else x
                            )
                        return df
                    except UnicodeDecodeError:
                        continue

                return pd.read_csv(file_path, engine='python', on_bad_lines='skip')
            
            elif file_path.endswith('.xlsx') | file_path.endswith('.xls'):
                try:
                    # use pandas to explicit engine close
                    with pd.ExcelFile(file_path, engine='openpyxl') as excel_file:
                        sheet_names = excel_file.sheet_names
                        print(f"Excel sheets found: {sheet_names}")
                    
                        for sheet in sheet_names:
                            try:
                                df = pd.read_excel(excel_file, sheet_name=sheet)
                                if not df.empty and len(df.columns) > 1:
                                    print(f"Found data in sheet: '{sheet}' with {len(df)} rows")
                                    # fixing dates parse format reading
                                    for col in df.columns:
                                        if pd.api.types.is_datetime64_any_dtype(df[col]):
                                            df[col] = df[col].strftime('%Y-%m-%d')
                                    return df
                            except Exception as e:
                                print(f"Could not read sheet '{sheet}': {e}")
                                continue
                        
                    logger.info(f"Trying openpyxl direct reading...")
                    from openpyxl import load_workbook
                
                    workbook = load_workbook(
                        file_path,
                        read_only=True,
                        data_only=True,
                        keep_links=False # this will prevent the file locking issue
                    )
                
                    try:
                        sheet = workbook.active
                        data = list(sheet.values)
                    
                        if not data:
                            print("No data found in workbook")
                            workbook.close()
                            return pd.DataFrame()
                    
                        cols = data[0]
                        df_data = data[1:] if len(data) > 1 else []
                        df = pd.DataFrame(df_data, columns=cols)
                        logger.info(f"Openpyxl loaded {len(df)} rows")
                    
                        workbook.close()
                        return df
                    except Exception as e:
                        workbook.close()
                        print(f"Openpyxl read failed: {e}")
                        raise
                
                except Exception as e:
                    print(f"Excel read failed: {e}")
                    try:
                        df = pd.read_excel(file_path, engine='openpyxl')
                        print(f"Fallback read successful: {len(df)} rows")
                        return df
                    except Exception as e:
                        print(f"Fallback also failed: {e}")
                        raise
        
        except Exception as e:
            print(f"Failed to read file: {e}")
            raise
                          
    def extract_transactions(self, file_path: str, user_id: int, document_id: int, column_mapping: dict) -> List[Transaction]:
        """Extract transactions from uploaded document with proper column mapping and template support"""
        try:
            # Read the file
            df = self._read_dataframe(file_path)
            print(f"📊 File loaded: {len(df)} rows, {len(df.columns)} columns")
            print(f"🔧 Column mapping: {column_mapping}")

            # convert column names to lowercase for matching
            df.columns = [str(col).strip().lower() for col in df.columns]

            # find mapped columns
            date_col = self._find_best_column_match(df.columns, column_mapping.get('date', 'date'))
            desc_col = self._find_best_column_match(df.columns, column_mapping.get('description', 'description'))
            amount_col = self._find_best_column_match(df.columns, column_mapping.get('amount', 'amount'))
            type_col = self._find_best_column_match(df.columns, column_mapping.get('type', 'type'))

            # detect document currency used
            detected_currency = self.detect_file_currency(file_path, df, amount_col)
            
            # build per row currency map from number_format, it handles where the number is plain integer but the number_format has currency symbol
            row_currency_map = {}
            if file_path.endswith(('.xlsx', '.xls')):
               
                original_amount_col = column_mapping.get('amount', 'amount')
                row_currency_map = self.excel_reader.read_with_currency_formats(file_path, original_amount_col)
                if row_currency_map:
                    # override the document detected currency with common per-cell currency
                    most_common = Counter(row_currency_map.values()).most_common(1)[0][0]
                    
                    detected_currency = most_common
                    logger.info(f"Per-row currency map built: {len(row_currency_map)} cells, predominant={most_common}")
            logger.info(f"Final detected currency: {detected_currency}")
            
            # update currency detector with the detected currency
            self.currency_detector = CurrencyDetector(base_currency='USD')
            self.detected_document_currency = detected_currency
            
            # update financial document record with detected currency
            try:
                doc = self.db.query(FinancialDocument).filter(FinancialDocument.id == document_id).first()
                if doc:
                    doc.document_currency = detected_currency
                    self.db.commit()
            except Exception as e:
                logger.warning(f"Could not update document currency: {e}")
            
            transactions_data = []
            successfull_rows = 0
            skipped_rows = 0

            # Process each row
            for index, row in df.iterrows():
                try:
                    # Skip if amount is empty
                    if pd.isna(row.get(amount_col)):
                        logger.info(f"⚠️ Row {index}: Skipping - Amount is empty")
                        skipped_rows += 1
                        continue

                    # parse amount
                    amount_raw = row[amount_col]
                    logger.info(f"🔍 Parsing amount from raw value: {amount_raw}")

                    # Check if the currency from excel number_format, if it is, prepend the symbol so it can be detected
                    if not isinstance(amount_raw, str) and isinstance(amount_raw, (int, float)):
                        
                        cell_currency = row_currency_map.get(index)
                        if cell_currency == 'IDR':
                            amount_str = f"Rp {amount_raw}"
                        elif cell_currency == 'USD':
                            amount_str = f"$ {amount_raw}"
                        else:
                            amount_str = str(amount_raw)
                            
                    elif not isinstance(amount_raw, str):
                        amount_str = str(amount_raw)
                    else:
                        amount_str = amount_raw
                        
                    logger.debug(f"Row {index}: amount_str for parsing = {amount_str!r}")

                    # currency detector on this part
                    try:
                        usd_amount, row_currency, currency_symbol = self.currency_detector.process_amount_string(amount_str)
                        
                        if not currency_symbol and hasattr(self, 'detected_document_currency'):
                            
                            doc_currency = self.detected_document_currency
                            if doc_currency != 'USD':
                                raw_amount = float(amount_raw) if isinstance(amount_raw, (int, float)) else float(str(amount_raw).replace(',', ''))
                                original_amount = raw_amount
                                usd_amount = self.currency_detector.convert_to_base_currency(raw_amount, doc_currency)
                                row_currency = doc_currency
                            else:
                                original_amount = usd_amount
                        elif row_currency != 'USD':
                            _, original_amt, _sym = self.currency_detector.detect_currency_from_string(amount_str)
                            original_amount = original_amt
                        else:
                            original_amount = usd_amount
                            
                        logger.debug(f"Row {index}: {amount_str} -> {usd_amount} USD (original: {original_amount} {detected_currency}, symbol: {currency_symbol})")
                        
                    except Exception as e:
                        logger.warning(f"Row {index}: Currency detection failed, trying fallback: {e}")
                    
                        # fallback to old method
                        amount_str_clean = amount_str.replace('$', '').replace('Rp', '').replace(',', '').replace(' ', '').strip()
                        is_negative = amount_str_clean.startswith('-') or amount_str_clean.startswith('(')
                        
                        if is_negative:
                            amount_str_clean = amount_str_clean.replace('-', '').replace('(', '').replace(')', '')
                            
                        usd_amount = pd.to_numeric(amount_str_clean, errors='coerce')
                        if pd.isna(usd_amount):
                            skipped_rows += 1
                            continue
                        
                        if is_negative:
                            usd_amount = -usd_amount
                            
                        detected_currency = 'USD'
                        original_amount = float(usd_amount)
                        currency_symbol = '$'

                    # parse date
                    date = self._parse_date(row.get(date_col, ''))
                    if pd.isna(date):
                        logger.info(f"⚠️ Row {index}: Skipping - Date parsing failed")
                        skipped_rows += 1
                        continue

                    # Get description and type
                    description = str(row.get(desc_col, '')).strip() or "Unknown Transactions"
                    type_value = str(row.get(type_col, '')).strip()

                    # Determine transaction type
                    transaction_type = self._determine_transaction_type(type_value, usd_amount, description)

                    # Caegorize transaction
                    category = categorize_transaction(description, self.db)

                    # Create transaction record
                    transaction_data = {
                        "document_id": document_id,
                        "user_id": user_id,
                        "date": date.to_pydatetime() if isinstance(date, pd.Timestamp) else date,
                        "description": description[:255],
                        "amount": float(usd_amount),
                        "type": transaction_type,
                        "category": category,
                        "month": date.strftime('%Y-%m') if hasattr(date, 'strftime') else str(date)[:7],
                        "created_at": datetime.now(),
                        "original_currency": detected_currency,
                        "original_amount": float(original_amount) if original_amount is not None else None,
                        "currency_symbol": currency_symbol if currency_symbol else None
                    }

                    transactions_data.append(transaction_data)
                    successfull_rows += 1

                    if successfull_rows % 100 == 0:
                        print(f"📝 Processed {successfull_rows} transactions...")

                except Exception as e:
                    print(f"⚠️ Skipping row {index}: {e}")
                    skipped_rows += 1
                    continue

            if transaction_data:
                self.db.bulk_insert_mappings(Transaction, transaction_data)
            self.db.commit()
            
            logger.info(f"Bulk inserted {successfull_rows} transactions, skipped {skipped_rows}")
            
            # return count instead of the ORM objects - bulk_insert_mappings does not
            return successfull_rows
        except Exception as e:
            self.db.rollback()
            print(f"❌ Error in extract_transactions: {e}")
            raise
        
    def _find_best_column_match(self, available_columns: List[str], target_column: str) -> str:
        """Find the best matching column name for template compatibility"""
        target_lower = target_column.lower()
        available_lower = [col.lower().strip() for col in available_columns]

        # if exact match exists
        if target_lower in available_columns:
            index = available_lower.index(target_lower)
            return available_columns[index]
        
        # Partial matches
        for i, col in enumerate(available_columns):
            col_lower = col.lower()
            if (target_lower in col_lower or 
                col_lower in target_lower or
                any(word in col_lower for word in target_lower.split()) and 
                any(word in target_lower for word in col_lower.split())):
                return available_columns[i]
            
        # Return first column as fallback
        return available_columns[0] if available_columns else target_column
    
    def _parse_date(self, date_value) -> pd.Timestamp:
        """Enhanced date parsing for various template formats"""
        if pd.isna(date_value):
            return pd.NaT

        try:
            # Handle various date formats
            if isinstance(date_value, (datetime, pd.Timestamp)):
                return pd.Timestamp(date_value)
            
            # Handle strings
            if isinstance(date_value, str):
                date_value = date_value.strip()
                # Remove time component if present
                date_value = date_value.split(' ')[0].split('T')[0]

            # handle DD/MM/YYYY format
            if '/' in date_value:
                try:
                    day, month, year = date_value.split('/')
                    # check if date is valid
                    if int(day) > 31 or int(month) > 12:
                        logger.warning(f"Invalid date format: {date_value}")
                        return pd.NaT
                
                    # fix invalid dates (like 31/06/2025 to 30/06/2025)
                    try:
                        return pd.Timestamp(int(year), int(month), int(day))
                    except ValueError:
                        # if day is invalid, set to last day of month
                        import calendar
                        last_day = calendar.monthrange(int(year), int(month))[1]
                        day = min(int(day), last_day)
                        return pd.Timestamp(int(year), int(month), day)
                except Exception:
                    pass
            return pd.to_datetime(date_value, errors='coerce', dayfirst=False)
        except:
            return pd.NaT
        
    def chunk_document(self, file_path: str, document_id: int, chunk_size: int = 1000) -> List[DocumentChunk]:
        """Chunk document for RAG processing"""
        try:
            print(f"📄 Starting advanced document chunking...")

            # Read file
            df = self._read_dataframe(file_path)

            if df.empty:
                print("⚠️ DataFrame is empty")
                return []
            
            print(f"📊 Processing {len(df)} rows for chunking")

            chunks = []

            # Method 1 : Group by semantic units (e.g., months, categories)
            text_chunks = self._create_semantic_chunks(df, chunk_size)

            # Method 2 : Create summary chunks
            summary_chunks = self._create_summary_chunks(df, chunk_size)

            # Combine strategies
            all_chunks = summary_chunks + text_chunks

            for chunk_index, chunk_text in enumerate(all_chunks):
                if len(chunk_text.strip()) < 50:
                    continue

                # Calculate embeddings-friendly metadata
                metadata = {
                    "chunk_size": len(chunk_text),
                    "document_id": document_id,
                    "chunk_index": chunk_index,
                    "total_chunks": len(all_chunks),
                    "source_file": os.path.basename(file_path),
                    "file_type": "excel" if file_path.endswith('.xlsx') else "csv",
                    "row_count": len(df),
                    "column_count": len(df.columns),
                    "chunk_strategy": "semantic" if chunk_index < len(summary_chunks) else "transaction",
                    "word_count": len(chunk_text.split()),
                    "char_count": len(chunk_text)
                }

                chunks.append(DocumentChunk(
                    document_id=document_id,
                    chunk_text=chunk_text,
                    chunk_index=chunk_index,
                    chunk_metadata=metadata
                ))                
                
            # add-all() sends one batched INSERT instead per one chunk that will took long time
            if chunks:
                self.db.add_all(chunks)
            self.db.commit()
            
            logger.info(f"Created {len(chunks)} document chunks for document {document_id}")
            return chunks
            
        except Exception as e:
            self.db.rollback()
            print(f"❌ Error in chunk_document: {e}")
            print(traceback.format_exc())
            raise

    def _create_semantic_chunks(self, df: pd.DataFrame, chunk_size: int) -> List[str]:
        """Create chunks based on semantic grouping"""
        chunks = []

        # Try to group by date (monthly chunks)
        date_cols = [col for col in df.columns if 'date' in col.lower()]
        if date_cols:
            date_col = date_cols[0]
            try:
                df['month'] = pd.to_datetime(df[date_col]).dt.to_period('M')

                for month, month_data in df.groupby('month'):
                    chunk_text = self._create_monthly_summary(month, month_data)
                    if len(chunk_text) > 100:
                        chunks.append(chunk_text)

            except:
                pass

        # Create transaction detail chunks (grouped for readability)
        transaction_chunks = self._create_transaction_chunks(df, chunk_size)
        chunks.extend(transaction_chunks)

        return chunks
    
    def _create_monthly_summary(self, month, month_data: pd.DataFrame) -> str:
        """Create monthly summary chunk"""
        text = f"📅 MONTHLY SUMMARY: {month}\n"
        text += "-" * 40 + "\n\n"

        text += f"📊 Transaction Count: {len(month_data)}\n"

        # Try to find amount column
        amount_cols = [col for col in month_data.columns if 'amount' in col.lower()]
        if amount_cols:
            amount_col = amount_cols[0]
            try:
                amounts = pd.to_numeric(month_data[amount_col], errors='coerce')
                if not amounts.isna().all():
                    total = amounts.sum()
                    avg = amounts.mean()
                    text += f"💰 Total Amount: ${total:,.2f}\n"
                    text += f"📈 Average Transaction: ${avg:,.2f}\n"
            except:
                pass

        # Top descriptions
        desc_cols = [col for col in month_data.columns if 'desc' in col.lower()]
        if desc_cols:
            desc_col = desc_cols[0]
            top_descs = month_data[desc_col].value_counts().head(3)
            if len(top_descs) > 0:
                text += "\n🔝 Top Transactions:\n"
                for desc, count in top_descs.items():
                    text += f"- {desc[:30]}{'...' if len(desc) > 30 else ''}: {count} times\n"
        
        return text
    
    def _create_summary_chunks(self, df: pd.DataFrame, chunk_size: int) -> List[str]:
        """Create summary/overview chunks"""
        chunks = []

        # Overall summary
        summary_text = "FINANCIAL DOCUMENT SUMMARY\n"
        summary_text += "=" * 40 + "\n\n"

        summary_text += f"📊 Document Overview:\n"
        summary_text += f"- Total transactions: {len(df)}\n"
        summary_text += f"- Columns available: {', '.join(df.columns[:5])}{'...' if len(df.columns) > 5 else ''}\n\n"

        # Column statistics
        summary_text += f"📈 Column Statistics:\n"
        for col in df.columns[:5]:
            non_null = df[col].count()
            unique = df[col].nunique()
            summary_text += f"- {col}: {non_null} values ({unique} unique)\n"

        chunks.append(summary_text)

        # Data type summary
        dtype_text = "🔍 Data Types and Patterns:\n"
        for col in df.columns[:5]:
            dtype = str(df[col].dtype)
            sample = str(df[col].iloc[0])[:50] if len(df) > 0 else "N/A"
            dtype_text += f"- {col}: {dtype} (sample: {sample}...)\n"

        chunks.append(dtype_text)
        return chunks
    
    def _create_transaction_chunks(self, df: pd.DataFrame, chunk_size: int) -> List[str]:
        """Create transaction detail chunks"""
        chunks = []
        current_chunk = "TRANSACTION DETAILS\n" + "=" * 40 + "\n\n"
        
        for index, row in df.iterrows():
            # Create transaction entry
            entry = f"Transaction {index+1}:\n"
            
            # Add key-value pairs for each column
            for col_idx, (col, value) in enumerate(row.items()):
                if pd.isna(value):
                    continue
                    
                value_str = str(value)
                if len(value_str) > 100:  # Truncate long values
                    value_str = value_str[:97] + "..."
                
                entry += f"  {col}: {value_str}\n"
            
            entry += "\n"
            
            # If adding this entry would exceed chunk size, start new chunk
            if len(current_chunk) + len(entry) > chunk_size and len(current_chunk) > 100:
                chunks.append(current_chunk)
                current_chunk = "TRANSACTION DETAILS (Continued)\n" + "=" * 40 + "\n\n"
            
            current_chunk += entry
        
        # Add the last chunk if it has content
        if len(current_chunk) > len("TRANSACTION DETAILS\n" + "=" * 40 + "\n\n"):
            chunks.append(current_chunk)
        
        return chunks
        
    def _determine_transaction_type(self, type_str: str, amount: float, description: str = "") -> str:
        """Determine if transaction is income oFr expense with multiple formats"""
        if pd.isna(type_str) or type_str == "":
            return "INCOME" if amount > 0 else "EXPENSE"
        
        type_clean = str(type_str).strip()
        type_lower = type_clean.lower()

        print(f"🔍 Type detection: '{type_str}' -> '{type_clean}' (lower: '{type_lower}'), amount: {amount}")

        # Comprehensive income keywords
        income_keywords = [
            'income', 'revenue', 'salary', 'deposit', 'credit', 
            'bonus', 'payment received', 'refund', 'interest',
            'incoming', 'payment from', 'transfer in', 'deposit', 'dividend'
        ]
        
        # Comprehensive expense keywords  
        expense_keywords = [
            'expense', 'payment', 'withdrawal', 'purchase', 'debit',
            'bill', 'fee', 'charge', 'payment sent', 'debit', 
            'negative', 'payment to', 'purchase at', 'withdrawal', 'atm'
        ]

        if any(keywords in type_lower for keywords in income_keywords):
            print(f"✅ Determined INCOME from description: {description}")
            return "INCOME"
            
        elif any(keywords in type_lower for keywords in expense_keywords):
            print(f"✅ Determined EXPENSE from description: {description}")
            return "EXPENSE"
        
         # Checking the description too if type is ambiguous
        desc_lower = description.lower()
        if any(keyword in desc_lower for keyword in ['salary', 'deposit', 'refund', 'interest', 'dividend', 'payment from']):
            return "INCOME"
        elif any(keyword in desc_lower for keyword in ['purchase', 'bill', 'fee', 'charge', 'payment', 'amazon', 'uber', 'netflix', 'rent', 'grocery', 'gas', 'electricity', 'water']):
            return "EXPENSE"
            
        else:
            result = 'INCOME' if amount > 0 else 'EXPENSE'
            print(f"🔍 Using amount-based fallback: {result}")
            return result
                
    def debug_excel_structure(self, file_path: str):
        """Debug method to examine Excel file structure"""
        print(f"\n🔍 DEBUG: Examining Excel file: {file_path}")
    
        try:
            # Check file exists
            if not os.path.exists(file_path):
                print("❌ File does not exist")
                return
        
            print(f"📊 File size: {os.path.getsize(file_path)} bytes")
        
            # Try pandas first
            try:
                excel_file = pd.ExcelFile(file_path)
                sheet_names = excel_file.sheet_names
                print(f"📄 Sheets found by pandas: {sheet_names}")
            
                for sheet in sheet_names:
                    try:
                        df = pd.read_excel(file_path, sheet_name=sheet, nrows=5)
                        print(f"\n📋 Sheet: '{sheet}'")
                        print(f"   Rows: {len(df)}, Columns: {list(df.columns)}")
                        print(f"   First 3 rows:")
                        for i in range(min(3, len(df))):
                            print(f"   Row {i}: {dict(df.iloc[i])}")
                    except Exception as e:
                        print(f"   ❌ Could not read sheet '{sheet}': {e}")
            except Exception as e:
                print(f"⚠️ Pandas failed: {e}")
        
            # Try openpyxl
            print("\n🔄 Trying openpyxl...")
            try:
                workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                print(f"📄 Sheets found by openpyxl: {workbook.sheetnames}")
            
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    print(f"\n📋 Sheet: '{sheet_name}'")
                    print(f"   Max row: {sheet.max_row}, Max column: {sheet.max_column}")
                
                # Read first few rows
                data = []
                for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    if i > 5:  # Only first 5 rows
                        break
                    data.append(row)
                
                if data:
                    print(f"   First {len(data)} rows:")
                    for i, row in enumerate(data):
                        print(f"   Row {i}: {row}")
                        
            except Exception as e:
                print(f"❌ Openpyxl failed: {e}")
            
        except Exception as e:
            print(f"❌ Debug error: {e}")
            print(f"🔍 Stack trace: {traceback.format_exc()}")

class ProcessingCancelledError(Exception):
    pass
        
class EnhancedDocumentService(DocumentService):
    def __init__(self, db: Session = None, user_currency: str = None):
        super().__init__(db, user_currency=user_currency)
        self.batch_processor = None
        self.cache = RedisCache()
        self.current_upload_id = None
        self.current_user_id = None   
        self.cancellation_checked = False
        self.db = db   
        self.enable_rate_limiting = True
        
    def set_progress(self, stage_index: int, custom_details: str = None, metadata: dict = None):
        """removed progress tracking"""
        pass
            
    def check_cancellation(self) -> bool:
        """Check if processing was cancelled via cancellation manager"""  
        if self.current_upload_id and self.current_user_id:
            from backend.api.routes.documents import upload_cancellation_manager
            cancelled = upload_cancellation_manager.is_cancelled(self.current_upload_id)
            if cancelled:
                self.cancellation_checked = True
                logger.info(f"Processing cancelled for upload {self.current_upload_id}")
            return cancelled
        return False
    
    def create_cancellation_check(self) -> Callable[[], bool]:
        """Create cancellation check function for batch processor"""
        def cancellation_check():
            return self.check_cancellation()
        return cancellation_check

    def _read_dataframe_chunked(self, file_path: str, chunk_size: int = 10000) -> pd.DataFrame:
        """Read a file into single frame."""
        if file_path.endswith('csv'):
            return pd.read_csv(file_path, low_memory=False)
        else:
            return self._read_dataframe(file_path)
        
    #@cached(category='document_processing', ttl=timedelta(hours=1))
    def process_document(self, file_path: str, user_id: int, filename: str, column_mapping: dict, cancellation_check=None) -> dict:
        """Optimized document processing with progress tracking along with cancellation progress"""
        # self track id
        self.current_user_id = user_id
        
        # cancellation check start point
        if cancellation_check is None:
            cancellation_check = self.create_cancellation_check()
        
        try:
            # initial stage  
            self.set_progress(0, f"Starting process for {filename}")
            
            # stage 1 : validating
            self.set_progress(1, "Checking file format...")
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File not found: {file_path}")
            
            # check cancellation
            if cancellation_check and cancellation_check():
                raise ProcessingCancelledError("Processing cancelled before validation")
            
            # stage 2 : reading file
            self.set_progress(2, f"Reading {filename}...")
            df = self._read_dataframe_chunked(file_path)
            
            # check cancellation
            if cancellation_check and cancellation_check():
                raise ProcessingCancelledError("Processing cancelled during file reading")
            
            # stage 3 : Parsing data
            self.set_progress(3, f"Parsing {len(df)} rows...")
            document = self.get_or_create_document(self.db, user_id, filename, file_path)
            
            # stage 4 : extracting transactions
            self.set_progress(4, "Extracting transactions from data...")
            transactions_data, extracted_data = self._prepare_transactions_batch(df, user_id, document.id, column_mapping, self.db, cancellation_check, file_path=file_path)
            
             # check cancellation
            if cancellation_check and cancellation_check():
                raise ProcessingCancelledError("Processing cancelled during transaction extraction")
            
            # stage 5 : Categorizing
            self.set_progress(5, f"Categorizing {len(transactions_data)} transactions...")
            
            # stage 6 : batch processing
            self.set_progress(6, "Processing transactions in batches...")
            
            # initialize batch processor with progress callback
            if not self.batch_processor:
                from backend.services.vector_search import VectorSearchService
                vector_service = VectorSearchService(self.db)
                
                self.batch_processor = BatchProcessor(
                    batch_size=500,
                    max_workers=4,
                    db=self.db,
                    vector_service=vector_service,
                    upload_id=self.current_upload_id,
                    user_id=self.current_user_id
                )
            
            # process transactions
            transaction_count = self.batch_processor.process_transactions_batch(
                transactions_data, cancellation_check
            )
            
            if cancellation_check is not None and cancellation_check():
                logger.info("Processing cancelled after batch processing")
                raise ProcessingCancelledError("Processing cancelled after batch processing")
            
            # stage 7 - 9 : Chunking + embeddings (will be skipped if the document only hold < 50 rows)
            EMBEDDING_ROW_THRESHOLD = 50
            chunks_data = []
            
            if len(df) >= EMBEDDING_ROW_THRESHOLD:
                self.set_progress(7, "Creating document chunks for search...")
                try:
                    chunks_data = self.batch_processor.chunk_documents_parallel(
                        df, document.id, cancellation_check=cancellation_check
                    )
                    if not chunks_data:
                        logger.warning("No chunks created from document")
                        chunks_data = []
                    logger.info(f"Created {len(chunks_data)} chunks from document")
                except Exception as e:
                    logger.error(f"Error during chunk creation: {e}")
                    chunks_data = []
                    
                if chunks_data:
                    self.set_progress(8, f"Generating embeddings for {len(chunks_data)} chunks...")
                    texts = [chunk['chunk_text'] for chunk in chunks_data]
                    embeddings = self.batch_processor.generate_embeddings_parallel(
                        texts, batch_size=50, max_workers=2, cancellation_check=cancellation_check
                    )
                    
                    if cancellation_check is not None and cancellation_check():
                        raise ProcessingCancelledError("Processing cancelled during embeddings")
                    
                    self.set_progress(9, "Storing embeddings in database...")
                    success = self.batch_processor.store_embeddings_parallel(
                        self.db, chunks_data, embeddings, cancellation_check=cancellation_check
                    )
                    if not success:
                        logger.warning("Some embeddings failed to store")
            else:
                logger.info(f"Skipping embeddings for small document ({len(df)} rows < {EMBEDDING_ROW_THRESHOLD}).")
                    
            # stage 10 : Finalizing
            self.set_progress(10, "Finalizing document processing...")
                
            # update document status
            document.processed = True
            document.processed_at = datetime.now()
            document.transaction_count = transaction_count
            self.db.commit()
                
            self.set_progress(12, f"Successfully processed {transaction_count} transactions")
                
            # cache metadata only
            file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0
            cache_key = f"doc:{user_id}:{file_size}"     
            
            self.cache.set('document_processing', cache_key, {
                "status": "success",
                "document_id": document.id,
                "transaction_count": transaction_count,
                "chunk_count": len(chunks_data) if chunks_data else 0,
                "processed_at": datetime.now().isoformat()
            }, ttl=timedelta(hours=24))
                
            logger.info(f"Document processing completed: {transaction_count} transactions")
                
            return {
                "document_id": document.id,
                "transaction_count": transaction_count,
                "chunk_count": len(chunks_data) if chunks_data else 0,
                "status": "success",
                "processing_time": datetime.now().isoformat()
            }
        
        except ProcessingCancelledError as e:
            logger.info(f"Document processing cancelled: {e}")
            raise
        
        except Exception as e:
            error_msg = str(e)[:200]
            logger.error(f"Document processing failed: {e}", exc_info=True)
            try:
                self.db.rollback()
            except:
                pass
            raise
    
    def store_extracted_transactions(self, extracted_data: List[Dict], db: Session, cancellation_check=None):
        """Store extracted transactions with parallel processing"""
        if not extracted_data:
            return 
        
        # add cancellation point at start
        if cancellation_check is not None and cancellation_check():
            logger.info("Processing cancelled before storing extracted transactions")
            raise Exception("Processing cancelled by user")
        
        batch_size = 100
        batches = [extracted_data[i:i + batch_size] for i in range(0, len(extracted_data), batch_size)]
        
        with ThreadPoolExecutor(max_workers=2) as executor:
            futures = []
            for batch_idx ,batch in enumerate(batches):
                if batch_idx % 5 == 0 and cancellation_check and cancellation_check():
                    logger.info(f"Processing cancelled during extracted transaction storage at batch {batch_idx}")
                    raise Exception("Processing cancelled by user")
                
                future = executor.submit(self.store_extracted_batch, db, batch, cancellation_check)
                futures.append(future)
                
            for future in as_completed(futures):
                try:
                    future.result()
                except Exception as e:
                    if "cancelled" in str(e).lower():
                        raise
                    logger.warning(f"Extracted transaction batch failed: {e}")
                        
    def store_extracted_batch(self, db: Session, batch: List[Dict], cancellation_check=None):
        """Store a batch of extracted transactions - uses a single IN() query for duplicate detection instead of one SELECT per record (N+1 fix)"""
        
        # cancellation point at the start
        if cancellation_check is not None and cancellation_check():
            logger.info("❌ Processing cancelled at start of extracted batch storage")
            raise Exception("Processing cancelled by user")
        
        if not batch:
            return
        
        # collect all raw_texts values from batch
        batch_raw_texts = {r.get('raw_texts', '') for r in batch}
        
        # one IN() query to find which raw_texts already exist instead of using loop
        try:
            existing_rows = db.query(ExtractedTransactions.raw_text).filter(
                ExtractedTransactions.raw_text.in_(batch_raw_texts)
            ).all()
            existing_texts = {row.raw_text for row in existing_rows}
        except Exception as e:
            logger.warning(f"Duplicate check query failed: {e} — inserting all records")
            existing_texts = set()
            
        # optional cancellation check after query
        if cancellation_check and cancellation_check():
            logger.info("Processing cancelled after duplicate check")
            raise Exception("Processing cancelled by user")
        
        # filter to only truly new records without DB
        new_records = [r for r in batch if r.get('raw_text', '') not in existing_texts]
        
        if not new_records:
            logger.debug(f"All {len(batch)} records in batch already exist — skipping")
            return
        
        # bulk insert all new records in one statement
        try:
            db.bulk_insert_mappings(ExtractedTransactions, new_records)
            db.commit()
            logger.debug(f"Bulk inserted {len(new_records)} new extracted transactions " f"({len(batch) - len(new_records)} duplicates skipped)")
        except Exception as e:
            db.rollback()
            logger.warning(f"Bulk insert failed: {e}")
            
    def get_or_create_document(self, db: Session, user_id: int, filename: str, file_path: str) -> FinancialDocument:
        """Get existing or create new document record"""
        document = db.query(FinancialDocument).filter(
            FinancialDocument.user_id == user_id,
            FinancialDocument.filename == filename
        ).order_by(FinancialDocument.id.desc()).first()
        
        if not document:
            file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0
            document = FinancialDocument(
                user_id=user_id,
                filename=filename,
                file_path=file_path,
                file_size=file_size,
                processed=False
            )
            db.add(document)
            db.commit()
            db.refresh(document)
        
        return document    
                    
    def _prepare_transactions_batch(self, df: pd.DataFrame, user_id: int, document_id: int, column_mapping: dict, db: Session, cancellation_check=None, file_path: str = None) -> Tuple[List[Dict], List[Dict]]:
        """Prepare transaction data dictionaries for batch insertion along with extracted chunks to save"""

        transactions_data = []
        extracted_data = []    
        
        # convert column names to lowercase for matching
        df.columns = [str(col).strip().lower() for col in df.columns]

        # find mapped columns
        date_col = self._find_best_column_match(df.columns, column_mapping.get('date', 'date'))
        desc_col = self._find_best_column_match(df.columns, column_mapping.get('description', 'description'))
        amount_col = self._find_best_column_match(df.columns, column_mapping.get('amount', 'amount'))
        type_col = self._find_best_column_match(df.columns, column_mapping.get('type', 'type'))
        
        # Build per-row currency map from Excel cell number_format.
        row_currency_map = {}
        if file_path and file_path.endswith(('.xlsx', '.xls')):
            original_amount_col = column_mapping.get('amount', 'Amount')
            row_currency_map = self.excel_reader.read_with_currency_formats(file_path, original_amount_col)
            if row_currency_map:
                most_common = Counter(row_currency_map.values()).most_common(1)[0][0]
                self.detected_document_currency = most_common
                logger.info(f"_prepare_transactions_batch: row_currency_map built, {len(row_currency_map)} cells → {most_common}")
        
        # cancellation point at start
        if cancellation_check is not None and cancellation_check():
            logger.info(f"Processing cancelled at start of transaction preparation")
            raise Exception("Process cancelled by user")
        
        # Process each row
        for index, row in df.iterrows():
            try:
                # periodic cancellation check every 10 rows
                if index % 10 == 0 and cancellation_check and cancellation_check():
                    logger.info(f"❌ Processing cancelled during row {index} preparation")
                    raise Exception("Processing cancelled by user")
                
                # Skip if amount is empty
                if pd.isna(row.get(amount_col)):
                    continue

                # Parse amount with currency detection
                amount_raw = row[amount_col]
                
                cell_currency = row_currency_map.get(index)
                usd_amount, detected_currency, original_amount, currency_symbol = self.parse_amount_with_currency(amount_raw, cell_currency=cell_currency)
                
                if pd.isna(usd_amount):
                    logger.debug(f"Row {index}: Skipping - amount parsing failed")
                    continue

                # Parse date 
                date = self._parse_date(row.get(date_col, ''))
                if pd.isna(date):
                    continue

                # Get description or type
                description = str(row.get(desc_col, '')).strip() or "Unknown Transactions"
                type_value = str(row.get(type_col, '')).strip()

                transaction_type = self._determine_transaction_type(type_value, usd_amount, description)
                # Categorize transaction
                category = categorize_transaction(description, db)

                # Ensure date is datetime object
                if isinstance(date, pd.Timestamp):
                    date_obj = date.to_pydatetime()
                    month_str = date_obj.strftime('%Y-%m')
                elif hasattr(date, 'strftime'):
                    date_obj = date
                    month_str = date.strftime('%Y-%m')
                else:
                    # Try to convert
                    try:
                        date_obj = datetime.fromisoformat(str(date))
                        month_str = date_obj.strftime('%Y-%m')
                    except:
                        print(f"⚠️ Row {index}: Could not format date, using current month")
                        date_obj = datetime.now()
                        month_str = date_obj.strftime('%Y-%m')
                        
                transaction_hash = hashlib.sha256(
                    f"{user_id}:{date_obj.strftime('%Y-%m-%d')}:{description[:100]}:{usd_amount:.2f}:{transaction_type}:{index}".encode()
                ).hexdigest()

                # create transaction dictionary (not an object ORM)
                transaction_data = {
                    "document_id": document_id,
                    "user_id": user_id,
                    "date": date.to_pydatetime() if isinstance(date, pd.Timestamp) else date,
                    "description": description[:255], 
                    "amount": float(usd_amount),
                    "type": transaction_type,
                    "category": category,
                    "month": month_str,
                    "transaction_hash": transaction_hash,
                    "created_at": datetime.now(),
                    "original_currency": detected_currency,
                    "original_amount": float(original_amount) if not pd.isna(original_amount) else float(usd_amount),
                    "currency_symbol": currency_symbol if currency_symbol else None
                }
                transactions_data.append(transaction_data)
                
                extracted_record = {
                    "user_id": user_id,
                    "document_id": document_id,
                    "date": date_obj,
                    "description": description[:255],
                    "amount": float(usd_amount),
                    "type": transaction_type,
                    "category": category,
                    "raw_text": str(row.to_dict()),
                    "metadata": {
                        "row_index": index,
                        "column_mapping": column_mapping,
                        "extraction_method": "document_processing",
                        "transaction_hash": transaction_hash
                    },
                    "year": date_obj.year,
                    "month": date_obj.month,
                    "is_processed": True,
                    "processed_at": datetime.now(),
                    "original_currency": detected_currency,
                    "original_amount": float(original_amount) if not pd.isna(original_amount) else float(usd_amount),
                    "currency_symbol": currency_symbol if currency_symbol else None
                }
                extracted_data.append(extracted_record)             
                
            except Exception as e:
                if "cancelled" in str(e).lower():
                    raise
                logger.info(f"⚠️ Skipping row {index} in batch preparation: {e}")
                continue
                  
        return transactions_data, extracted_data
    
    def parse_amount_with_currency(self, amount_raw, cell_currency: str = None) -> Tuple[float, str, float, str]:
        """Parse amount with currencies detector"""
        if pd.isna(amount_raw):
            return (float('nan'), 'USD', float('nan'), '')
    
        try:
            if cell_currency and isinstance(amount_raw, (int, float)):
                original_amount = float(amount_raw)
                usd_amount = self.currency_detector.convert_to_base_currency(original_amount, cell_currency)
                currency_symbol = 'Rp' if cell_currency == 'IDR' else '$' if cell_currency == 'USD' else ''
                logger.debug(f"Parsed via cell format: {amount_raw} {cell_currency} -> ${usd_amount:.4f} USD")
                return (usd_amount, cell_currency, original_amount, currency_symbol)
            
            # convert to string
            if not isinstance(amount_raw, str):
                amount_str = str(amount_raw)
            else:
                amount_str = amount_raw
                
            usd_amount, detected_currency, currency_symbol = self.currency_detector.process_amount_string(amount_str)
            _, original_amount, _ = self.currency_detector.detect_currency_from_string(amount_str)
            
            logger.debug(f"Parsed: '{amount_str}' -> USD: ${usd_amount:.2f}, Original: {original_amount} {detected_currency}, Symbol: {currency_symbol}")
            return (usd_amount, detected_currency, original_amount, currency_symbol)
        
        except Exception as e:
            logger.error(f"Currency-aware parsing failed for '{amount_raw}': {e}")
            return (float('nan'), 'USD', float('nan'), '')